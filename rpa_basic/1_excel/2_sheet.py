from tkinter import W
from openpyxl import Workbook
wb =Workbook()
ws =wb.create_sheet() # 새로운 sheet 생성 
ws.title="MySheet" # sheet 이름변경 
ws.sheet_properties.tabColor="ff00ff" 
wb.save("sample.xlsx")
# sheet,MySheet,YourSheet
ws1=wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번쨰 index에 생성 


new_ws = wb["NewSheet"] # dict 형태로 sheet에 접근 가능 

print(wb.sheetnames) # 모든 시트 이름 확인 


# sheet 복사 
new_ws["A1"] ="Test"
target = wb.copy_worksheet(new_ws)
target.title="copied sheet"


wb.save("sample.xlsx")

